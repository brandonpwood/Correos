import time

from openpyxl import Workbook, load_workbook
from docx import Document

class Invoicer():
    def cast_data(self, project, date):
        '''Reads data from project and returns a dictionary with formatting
        for an invoice to edit with.
        '''
        return {
            '\PO': str(project.data['po_number']),
            '\REP': str(project.data['sales_rep']),
            '\TOTAL': str(project.data['drawdown']),
            '\DOLLARTOTAL': '$' + str(project.data['drawdown']),
            '\DATE': date,
            '\TAXID': project.data['taxid'],
            '\REP': project.data['sales_rep'],
            '\REPLINE1': project.data['rep_line_1'],
            '\REPLINE2': project.data['rep_line_2'],
            '\REPEMAIL': project.data['rep_email'],
            '\CLIENT': project.data['client'],
            '\CLIENTLINE1': project.data['client_line_1'],
            '\CLIENTLINE2': project.data['client_line_2'],
            '\CLIENTEMAIL': project.data['client_email'],
            '\VENDOR': project.data['vendor']
        }

    def generate_invoices(self, template, project, date):
        '''Generates an invoice for a project from the given template file.
        '''
        document = Document(template)
        data = self.cast_data(project, date)

        # Convert Tables.
        for table in document.tables:
            for col in table.columns:
                for cell in col.cells:
                    if cell.text in data.keys():
                        cell.text = data[cell.text]

        document.save(project.data['reference'] + '_invoice.docx')
        print('document saved as:', project.data['reference'] + '_invoice.docx')

    def update_budget_tracker(self, tracker, month):
        '''Given a budget tracker and the current projects, update drawdowns
        for this month.
        '''
        wb = load_workbook(tracker)
        ws = wb.active
        print()

        # Find last column.
        last_col = 0
        for num, col in enumerate(ws.iter_cols(min_row = 0, min_col = 0)):
            if not col[0].value:
                last_col = num
                break
        # If it doesn't find an empty column.
        if last_col == 0:
            last_col = len(ws.iter_cols(min_row = 0, min_col = 0))

        # New month is one beyond the last
        last_col += 1

        # Add title
        ws.cell(row = 1, column = last_col, value = month)

        # Updates. Loop through and find row that corresponds to projects.
        overdrawn = []
        for project in self.projects:
            row_number = 1 # iterable object from worksheet is immutable.
            print('Updating ' + project.data['reference'])
            for row in ws:
                if row[2].value == project.data['reference']:
                    # Subtract.
                    ws.cell(row = row_number, column = 2, value = row[5].internal_value - project.data['drawdown'])

                    # Check if overdrawn.
                    if float(row[5].value) - project.data['drawdown'] <= 0:
                        overdrawn.append(project.data['reference'])

                    # Append to month's drawdown.
                    ws.cell(row = row_number, column = last_col, value = project.data['drawdown'])
                    break
                row_number += 1

        wb.save('Updated Budget Tracker for ' + month + '.xlsx')

        print('Updates complete. If a project budget is overdrawn, I will list it below:')
        if overdrawn != []:
            for project_name in overdrawn:
                print(project_name)
        else:
            print('None found! Fantastic!')

    def read_projects(self, name):
        ''' Read projects from template or old report file.
        '''
        wb = load_workbook(name)
        ws = wb.active

        # Collect field info, split into keys for employees and projects.
        fields =  [x.value for x in ws[1]]
        project_fields = fields[:fields.index('name')]
        employee_fields = fields[fields.index('name'):]

        # Generate Projects
        project_data_groups = []
        employee_data_groups = []
        group = []

        for row in ws:
            if row[0] != ws['A1']:
                if row[0].value != None:
                    data = {}
                    for num, field in enumerate(project_fields):
                        data[field] = row[num].value
                    # Ignore first project when no employees have been seen yet.
                    # Employee groups are always one behind the projects.
                    if len(project_data_groups) != 0:
                        employee_data_groups.append(group)
                        group = []
                    project_data_groups.append(data)

                # Add employee to group
                employee = {}
                for num, field in enumerate(employee_fields):
                    employee[field] = row[num + len(project_fields)].value
                group.append(employee)
        employee_data_groups.append(group)

        for group, emps in zip(project_data_groups, employee_data_groups):
            self.projects.append(self.project(group, emps))
        print('Projects Added!')
    def generate_very_simple(self, users):
        '''Generate simple object without projects.
        '''
        wb = Workbook
        ws = wb.active

        ws.cell(row = 1, column = 1, value = 'Names')
        ws.cell(row = 1, column = 2, value = 'Days')

        row_number = 0

        for user in users:
            ws.cell(row = row_number, column = 1, value  = user[0])
            ws.cell(row = row_number, column = 2, value  = user[1])

            row_number += 1

        wb.save(time + '.xlsx')


    def generate_simple(self, name):
        ''' Stores current employee data in a spreadsheet for updating.
        '''
        wb = Workbook()
        ws = wb.active

        # Populate titles
        for num, title in enumerate(self.projects[-1].employees[-1].data):
            ws.cell(row = 1, column = num + 1, value = title)

        # Populate data.
        row_number = 2
        for project in self.projects:
            for employee in project.employees:
                for num, datum in enumerate(employee.data.values()):
                    ws.cell(row = row_number, column = num + 1, value = datum)

                row_number += 1

        wb.save(name)

    def read_simple(self, name):
        '''Update project and employee data from a simple file.
        '''
        wb = load_workbook(name)
        ws = wb.active

        # Get titles
        titles = [x.value for x in ws[1]]

        for project in self.projects:
            for employee in project.employees:
                for row in ws:
                    if row[titles.index('name')] == employee.name:
                        for title, val  in zip(titles, row):
                            employee.data[title] = val
                        break
        # Update expenses
        for project in self.projects:
            project.update_expenses()

    def generate_report(self, file_name):
        ''' Saves active projects to a template file of the given name.
        '''
        wb = Workbook()
        ws = wb.active

        # Populate titles.
        length = len(self.projects[-1].data)

        for num, title in enumerate(self.projects[-1].data):
            ws.cell(row = 1, column = num + 1, value = title)
        for num, title in enumerate(self.projects[-1].employees[-1].data):
            ws.cell(row = 1, column = length + 1 + num, value = title)

        # Populate projects.
        row_number = 2
        for project in self.projects:
            length = len(project.data.values())
            for num, value in enumerate(project.data.values()):
                ws.cell(row = row_number, column = num + 1, value = value)
            for employee in project.employees:
                for num, attr in enumerate(employee.data.values()):
                    ws.cell(row =  row_number, column = length + 1 + num, value = attr)
                row_number += 1
        wb.save(file_name)
        print('Report Generated!')

    class project:
        ''' Project Class. Holds general information on projects as well as
        employee data. Must be initialized with a dictionary and a list of
        dictionaries for the employees.
        '''
        def __init__(self, data, employees):
            # Save info
            self.data = data

            self.data['budget_left'] = data['budget']

            # Make and add all the employees
            self.employees = []
            for data in employees:
                self.employees.append(self.employee(data))

            # Update as best as possible
            self.update_expenses()

        def update_employee_expenses(self):
            ''' Update expenses(pay and cost) for every employee on the project.
            '''
            for employee in self.employees:
                employee.update_expenses()

        def update_expenses(self):
            ''' Update Employee expenses then update the drawdown, budget, and
            profit for the project.
            '''
            self.update_employee_expenses()

            total_cost = sum([x.cost for x in self.employees])
            total_pay = sum([x.pay for x in self.employees])

            self.data['drawdown'] = total_cost
            self.data['profit'] = total_cost - total_pay
            self.data['budget_left'] -= total_cost

        class employee:
            ''' Holds data on an employee. Initialized with a dictionary. Must
            have name, pay_type, and a cost_rate and pay_rate, whether
            that be daily or a salary.

            The cost_rate denotes the rate that we charge the client per unit of
            time, and the pay is what is earned by the employee in that time.
            Both are in dollars and are floats.
            '''
            def __init__(self, data):
                self.data = data

                # Make it easier to access name
                self.name = data['name']

                # Adjusted when days_worked is provided.
                self.cost = 0
                self.pay = 0
                self.days_worked = 0

                if self.data['pay_type'] is 'salary':
                    self.cost = self.data['cost_rate']
                    self.pay =  self.data['cost']

            def update_expenses(self):
                '''Updates pay, depending on the pay_type.
                '''
                if self.data['pay_type'] is 'daily':
                    self.cost = self.days_worked * self.data['cost_rate']
                    self.pay =  self.days_worked * self.data['pay_rate']

                if self.data['pay_type'] is 'hours':
                    self.cost = self.days_worked * self.cost_rate
                    self.pay =  self.days_worked * self.pay_rate

                if self.data['pay_type'] is 'salary':
                    self.cost = self.cost_rate
                    self.pay =  self.pay_rate
